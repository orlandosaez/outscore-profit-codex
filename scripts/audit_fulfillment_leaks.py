from __future__ import annotations

import argparse
import csv
import shlex
import subprocess
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ASSIGNMENTS_XLSX = ROOT / "docs/data-references/client-staff-assignments.xlsx"
ANCHOR_SERVICES_CSV = ROOT / "docs/data-references/anchor services.csv"
OUTPUT_CSV = Path("/tmp/audit_fulfillment_leaks_enriched.csv")
DEFAULT_SSH_TARGET = "root@104.225.220.36"
DEFAULT_SSH_PORT = "2222"
DEFAULT_REMOTE_ENV = "/opt/agents/outscore_profit/.env"

INTERNAL_NAME_MARKERS = (
    "saez",
    "sbc accounting and tax",
    "outscore",
)


@dataclass
class AssignmentRecord:
    client_name: str
    group: str | None
    services: list[str] = field(default_factory=list)
    related_entities: list[str] = field(default_factory=list)
    staff_primary: str = ""
    staff_reviewer: str = ""


@dataclass
class BillingSummary:
    billed_members: list[str] = field(default_factory=list)
    invoice_count: int = 0
    total_invoiced: float = 0.0


@dataclass(frozen=True)
class ServicePrice:
    service_name: str
    fc_tag: str
    annualized_price: float


def normalize_name(value: str) -> str:
    return " ".join(value.strip().lower().split())


def split_tokens(value: object) -> list[str]:
    if value is None:
        return []
    return [token.strip() for token in str(value).split(";") if token and token.strip()]


def parse_group_and_services(group_service_text: object) -> tuple[str | None, list[str], list[str]]:
    tokens = split_tokens(group_service_text)
    services = [token for token in tokens if token.upper().startswith("S ")]
    non_service_tokens = [token for token in tokens if not token.upper().startswith("S ")]
    group = non_service_tokens[0] if non_service_tokens else None
    related_entities = non_service_tokens[1:] if group else []
    return group, services, related_entities


def merge_text(existing: str, new_value: object) -> str:
    new_text = "" if new_value is None else str(new_value).strip()
    if not new_text:
        return existing
    if not existing:
        return new_text
    parts = [part.strip() for part in existing.split(";") if part.strip()]
    if new_text not in parts:
        parts.append(new_text)
    return "; ".join(parts)


def parse_assignments(path: Path = ASSIGNMENTS_XLSX) -> tuple[dict[str, AssignmentRecord], dict[str, list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    assignments: dict[str, AssignmentRecord] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        client_name = "" if row[0] is None else str(row[0]).strip()
        if not client_name:
            continue
        group, services, related_entities = parse_group_and_services(row[1])
        key = normalize_name(client_name)
        record = assignments.get(key)
        if record is None:
            assignments[key] = AssignmentRecord(
                client_name=client_name,
                group=group,
                services=[],
                related_entities=[],
                staff_primary="" if row[2] is None else str(row[2]).strip(),
                staff_reviewer="" if row[3] is None else str(row[3]).strip(),
            )
            record = assignments[key]
        elif record.group is None and group is not None:
            record.group = group
            record.staff_primary = merge_text(record.staff_primary, row[2])
            record.staff_reviewer = merge_text(record.staff_reviewer, row[3])
        else:
            record.staff_primary = merge_text(record.staff_primary, row[2])
            record.staff_reviewer = merge_text(record.staff_reviewer, row[3])

        for service in services:
            if service not in record.services:
                record.services.append(service)
        for related_entity in related_entities:
            if related_entity not in record.related_entities:
                record.related_entities.append(related_entity)

    groups: dict[str, list[str]] = {}
    for record in assignments.values():
        if record.group:
            groups.setdefault(record.group, []).append(record.client_name)
    for members in groups.values():
        members.sort(key=str.lower)
    return assignments, groups


def parse_money(value: object) -> float:
    if value is None:
        return 0.0
    text = str(value).replace("$", "").replace(",", "").strip()
    if not text:
        return 0.0
    return float(text)


def annualize_price(price: float, billing_occurrence: str) -> float:
    occurrence = billing_occurrence.strip().lower()
    if occurrence == "monthly":
        return price * 12
    if occurrence == "quarterly":
        return price * 4
    return price


def load_service_prices_by_tag(path: Path = ANCHOR_SERVICES_CSV) -> dict[str, list[ServicePrice]]:
    prices: dict[str, list[ServicePrice]] = {}
    with path.open(encoding="utf-8-sig", newline="") as file:
        for row in csv.DictReader(file):
            if row.get("Type") != "Service":
                continue
            tag = (row.get("Tag") or "").strip()
            if not tag:
                continue
            price = parse_money(row.get("Price"))
            prices.setdefault(tag, []).append(
                ServicePrice(
                    service_name=(row.get("Name") or "").strip(),
                    fc_tag=tag,
                    annualized_price=annualize_price(
                        price,
                        row.get("Billing Occurrence") or "",
                    ),
                )
            )
    return prices


def estimate_annual_revenue(
    services: list[str],
    service_prices_by_tag: dict[str, list[ServicePrice]],
) -> float:
    total = 0.0
    for service_tag in services:
        prices = service_prices_by_tag.get(service_tag, [])
        if not prices:
            continue
        # Umbrella tags such as S BILL can map to multiple pass-through services.
        # Use the highest annualized price so the audit does not understate risk.
        total += max(price.annualized_price for price in prices)
    return total


def run_psql_csv(sql: str, *, ssh_target: str, ssh_port: str, remote_env: str) -> list[dict[str, str]]:
    remote_command = (
        f"set -a; . {shlex.quote(remote_env)}; set +a; "
        f"psql \"$SUPABASE_DB_URL\" --csv -P footer=off -c {shlex.quote(sql)}"
    )
    result = subprocess.run(
        ["ssh", "-p", ssh_port, ssh_target, remote_command],
        check=True,
        text=True,
        capture_output=True,
    )
    output = result.stdout.strip()
    if not output:
        return []
    return list(csv.DictReader(StringIO(output)))


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def fetch_leak_candidates(*, ssh_target: str, ssh_port: str, remote_env: str) -> list[dict[str, str]]:
    sql = """
WITH fc_clients_with_completed_work AS (
  SELECT DISTINCT
    fc.fc_client_id,
    fc.name as fc_client_name
  FROM profit_fc_clients fc
  JOIN profit_fc_tasks ft ON ft.fc_client_id = fc.fc_client_id
  WHERE ft.is_completed = true
    AND ft.completed_at >= (current_date - interval '180 days')
),
fc_clients_with_anchor_invoices AS (
  SELECT DISTINCT
    match.fc_client_id
  FROM profit_fc_client_anchor_matches match
  JOIN profit_anchor_invoices ai ON ai.anchor_relationship_id = match.anchor_relationship_id
  WHERE ai.issue_date >= (current_date - interval '180 days')
)
SELECT
  fcc.fc_client_id,
  fcc.fc_client_name
FROM fc_clients_with_completed_work fcc
LEFT JOIN fc_clients_with_anchor_invoices fci ON fci.fc_client_id = fcc.fc_client_id
WHERE fci.fc_client_id IS NULL
ORDER BY fcc.fc_client_name;
"""
    return run_psql_csv(sql, ssh_target=ssh_target, ssh_port=ssh_port, remote_env=remote_env)


def fetch_anchor_matches_for_client_names(
    client_names: list[str],
    *,
    ssh_target: str,
    ssh_port: str,
    remote_env: str,
) -> dict[str, list[str]]:
    if not client_names:
        return {}
    names = ", ".join(sql_string(name) for name in client_names)
    sql = f"""
WITH input_names AS (
  SELECT
    name,
    profit_normalize_client_name(name) as normalized_name
  FROM unnest(ARRAY[{names}]::text[]) as input(name)
)
SELECT
  input.name as fc_client_name,
  aa.client_business_name as anchor_client_name
FROM input_names input
JOIN profit_anchor_agreements aa
  ON profit_normalize_client_name(aa.client_business_name) = input.normalized_name
ORDER BY input.name, aa.client_business_name;
"""
    rows = run_psql_csv(sql, ssh_target=ssh_target, ssh_port=ssh_port, remote_env=remote_env)
    matches: dict[str, list[str]] = {}
    for row in rows:
        matches.setdefault(normalize_name(row["fc_client_name"]), []).append(
            row["anchor_client_name"]
        )
    return matches


def fetch_group_billing(
    group_members: list[str],
    *,
    ssh_target: str,
    ssh_port: str,
    remote_env: str,
) -> BillingSummary:
    if not group_members:
        return BillingSummary()
    member_list = ", ".join(sql_string(member) for member in group_members)
    sql = f"""
WITH group_member_names AS (
  SELECT profit_normalize_client_name(name) as normalized_name
  FROM unnest(ARRAY[{member_list}]::text[]) as input(name)
)
SELECT
  aa.client_business_name,
  count(ai.anchor_invoice_id) as invoice_count_180d,
  coalesce(sum(ai.total_amount), 0) as invoiced_180d
FROM profit_anchor_agreements aa
LEFT JOIN profit_anchor_invoices ai
  ON ai.anchor_relationship_id = aa.anchor_relationship_id
  AND ai.issue_date >= (now() - interval '180 days')
WHERE profit_normalize_client_name(aa.client_business_name) IN (
  SELECT normalized_name
  FROM group_member_names
)
GROUP BY 1
ORDER BY 1;
"""
    rows = run_psql_csv(sql, ssh_target=ssh_target, ssh_port=ssh_port, remote_env=remote_env)
    billed_members: list[str] = []
    invoice_count = 0
    total_invoiced = 0.0
    for row in rows:
        count = int(row["invoice_count_180d"] or 0)
        amount = float(row["invoiced_180d"] or 0)
        if count > 0:
            billed_members.append(row["client_business_name"])
        invoice_count += count
        total_invoiced += amount
    return BillingSummary(
        billed_members=billed_members,
        invoice_count=invoice_count,
        total_invoiced=total_invoiced,
    )


def is_internal_account(client_name: str) -> bool:
    normalized = normalize_name(client_name)
    return any(marker in normalized for marker in INTERNAL_NAME_MARKERS)


def classify_candidate(
    *,
    fc_client_name: str,
    assignment: AssignmentRecord | None,
    billing: BillingSummary,
    anchor_match_exists: bool,
) -> str:
    if is_internal_account(fc_client_name):
        return "internal_account"
    if assignment is None:
        return "not_in_assignments_file"
    if assignment.group and billing.invoice_count > 0:
        return "consolidated_via_group_billed"
    if assignment.group and billing.invoice_count == 0:
        return "consolidated_via_group_unbilled"
    if assignment.group is None and anchor_match_exists:
        return "name_mismatch_anchor_exists"
    if assignment.group is None and not anchor_match_exists:
        return "standalone_no_anchor"
    return "unclassified"


def needs_manual_review(classification: str) -> bool:
    return classification in {
        "consolidated_via_group_unbilled",
        "standalone_no_anchor",
        "name_mismatch_anchor_exists",
        "not_in_assignments_file",
        "unclassified",
    }


def enrich_candidates(
    candidates: list[dict[str, str]],
    assignments: dict[str, AssignmentRecord],
    groups: dict[str, list[str]],
    anchor_matches_by_client_name: dict[str, list[str]],
    service_prices_by_tag: dict[str, list[ServicePrice]],
    *,
    ssh_target: str,
    ssh_port: str,
    remote_env: str,
) -> list[dict[str, object]]:
    billing_by_group: dict[str, BillingSummary] = {}
    enriched: list[dict[str, object]] = []
    for candidate in candidates:
        fc_client_name = candidate["fc_client_name"]
        assignment = assignments.get(normalize_name(fc_client_name))
        group_name = assignment.group if assignment else None
        group_members = groups.get(group_name, []) if group_name else []
        if group_name and group_name not in billing_by_group:
            billing_by_group[group_name] = fetch_group_billing(
                group_members,
                ssh_target=ssh_target,
                ssh_port=ssh_port,
                remote_env=remote_env,
            )
        billing = billing_by_group.get(group_name or "", BillingSummary())
        classification = classify_candidate(
            fc_client_name=fc_client_name,
            assignment=assignment,
            billing=billing,
            anchor_match_exists=bool(anchor_matches_by_client_name.get(normalize_name(fc_client_name))),
        )
        estimated_annual_revenue = estimate_annual_revenue(
            assignment.services if assignment else [],
            service_prices_by_tag,
        )
        enriched.append(
            {
                "fc_client_id": candidate["fc_client_id"],
                "fc_client_name": fc_client_name,
                "group_name": group_name or "",
                "group_members": "; ".join(group_members),
                "group_members_with_180d_invoices": "; ".join(billing.billed_members),
                "group_members_invoice_count": billing.invoice_count,
                "group_members_total_invoiced_180d": f"{billing.total_invoiced:.2f}",
                "staff_primary": assignment.staff_primary if assignment else "",
                "staff_reviewer": assignment.staff_reviewer if assignment else "",
                "estimated_annual_revenue": f"{estimated_annual_revenue:.2f}",
                "suggested_classification": classification,
                "needs_manual_review": "yes" if needs_manual_review(classification) else "no",
            }
        )
    return enriched


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = [
        "fc_client_id",
        "fc_client_name",
        "group_name",
        "group_members",
        "group_members_with_180d_invoices",
        "group_members_invoice_count",
        "group_members_total_invoiced_180d",
        "staff_primary",
        "staff_reviewer",
        "estimated_annual_revenue",
        "suggested_classification",
        "needs_manual_review",
    ]
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def print_summary(rows: list[dict[str, object]]) -> None:
    counts = Counter(str(row["suggested_classification"]) for row in rows)
    print("Fulfillment leak audit enrichment")
    print(f"Generated: {date.today().isoformat()}")
    print("Source note: docs/data-references/client-staff-assignments.xlsx is a transitional FC tag snapshot, not the canonical source.")
    print("")
    print("Counts by suggested classification:")
    for classification, count in sorted(counts.items()):
        print(f"  {classification}: {count}")
    manual_review_count = sum(1 for row in rows if row["needs_manual_review"] == "yes")
    risk_buckets = [
        "consolidated_via_group_unbilled",
        "standalone_no_anchor",
        "name_mismatch_anchor_exists",
    ]
    exposure_by_bucket = {
        bucket: sum(
            float(row["estimated_annual_revenue"])
            for row in rows
            if row["suggested_classification"] == bucket
        )
        for bucket in risk_buckets
    }
    print("")
    print(f"Manual review rows: {manual_review_count}")
    print("Estimated annual revenue at risk:")
    for bucket in risk_buckets:
        print(f"  {bucket}: ${exposure_by_bucket[bucket]:,.2f}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Enrich FC fulfillment leak candidates with transitional assignment snapshot group/service data."
        )
    )
    parser.add_argument("--assignments", type=Path, default=ASSIGNMENTS_XLSX)
    parser.add_argument("--output", type=Path, default=OUTPUT_CSV)
    parser.add_argument("--ssh-target", default=DEFAULT_SSH_TARGET)
    parser.add_argument("--ssh-port", default=DEFAULT_SSH_PORT)
    parser.add_argument("--remote-env", default=DEFAULT_REMOTE_ENV)
    args = parser.parse_args()

    assignments, groups = parse_assignments(args.assignments)
    service_prices_by_tag = load_service_prices_by_tag()
    candidates = fetch_leak_candidates(
        ssh_target=args.ssh_target,
        ssh_port=args.ssh_port,
        remote_env=args.remote_env,
    )
    anchor_matches_by_client_name = fetch_anchor_matches_for_client_names(
        [candidate["fc_client_name"] for candidate in candidates],
        ssh_target=args.ssh_target,
        ssh_port=args.ssh_port,
        remote_env=args.remote_env,
    )
    rows = enrich_candidates(
        candidates,
        assignments,
        groups,
        anchor_matches_by_client_name,
        service_prices_by_tag,
        ssh_target=args.ssh_target,
        ssh_port=args.ssh_port,
        remote_env=args.remote_env,
    )
    write_csv(args.output, rows)
    print_summary(rows)
    print("")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
