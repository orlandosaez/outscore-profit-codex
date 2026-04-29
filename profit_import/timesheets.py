from __future__ import annotations

import csv
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl


STAFF_RATES: dict[str, float] = {
    "Beth": 45.0,
    "Julie": 30.0,
    "Laura": 60.0,
    "Noelle": 55.0,
    "Wama": 16.0,
}

WAMA_VALID_AFTER = date(2025, 9, 15)

ADMIN_PATTERNS = (
    "admin",
    "internal",
    "outscore",
    "sbc",
    "meeting",
)


@dataclass(frozen=True)
class TimeEntry:
    staff_name: str
    entry_date: date
    client_raw: str
    task_raw: str
    hours: float
    hourly_rate: float
    labor_cost: float
    service_type: str
    is_admin: bool
    source_file: str
    source_sheet: str
    source_row: int
    paid_marker: str | None = None


def parse_timesheet_folder(folder: Path | str) -> list[TimeEntry]:
    folder_path = Path(folder)
    entries: list[TimeEntry] = []

    for path in sorted(folder_path.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        entries.extend(parse_timesheet_file(path))

    return entries


def parse_timesheet_file(path: Path | str) -> list[TimeEntry]:
    file_path = Path(path)
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    first_sheet = workbook.worksheets[0]
    first_row = _normalized_row(next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    lower_name = file_path.name.lower()

    if first_row[:4] == ["client", "staff", "hrs", "date"]:
        return _parse_laura(file_path, first_sheet)
    if first_row[:4] == ["customer name", "note / task", "hours", "date"]:
        return _parse_wama(file_path, first_sheet)
    if "julie" in lower_name:
        return _parse_julie(file_path, first_sheet)
    if "beth" in lower_name:
        return _parse_beth(file_path, first_sheet)

    raise ValueError(f"Unsupported timesheet format: {file_path}")


def write_time_entries_csv(entries: Iterable[TimeEntry], output_path: Path | str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [asdict(entry) for entry in entries]

    fieldnames = [
        "staff_name",
        "entry_date",
        "client_raw",
        "task_raw",
        "hours",
        "hourly_rate",
        "labor_cost",
        "service_type",
        "is_admin",
        "source_file",
        "source_sheet",
        "source_row",
        "paid_marker",
    ]

    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def summarize_time_entries(entries: Iterable[TimeEntry]) -> dict[str, object]:
    rows = list(entries)
    staff_summary: dict[str, dict[str, float | int]] = {}

    for entry in rows:
        staff = staff_summary.setdefault(
            entry.staff_name,
            {
                "entry_count": 0,
                "hours": 0.0,
                "labor_cost": 0.0,
                "admin_hours": 0.0,
                "client_allocated_hours": 0.0,
            },
        )
        staff["entry_count"] += 1
        staff["hours"] = round(float(staff["hours"]) + entry.hours, 2)
        staff["labor_cost"] = round(float(staff["labor_cost"]) + entry.labor_cost, 2)
        if entry.is_admin:
            staff["admin_hours"] = round(float(staff["admin_hours"]) + entry.hours, 2)
        else:
            staff["client_allocated_hours"] = round(float(staff["client_allocated_hours"]) + entry.hours, 2)

    return {
        "entry_count": len(rows),
        "total_hours": round(sum(entry.hours for entry in rows), 2),
        "labor_cost": round(sum(entry.labor_cost for entry in rows), 2),
        "admin_hours": round(sum(entry.hours for entry in rows if entry.is_admin), 2),
        "client_allocated_hours": round(sum(entry.hours for entry in rows if not entry.is_admin), 2),
        "staff": staff_summary,
    }


def infer_service_type(task_raw: object, client_raw: object) -> str:
    text = f"{_clean_text(task_raw)} {_clean_text(client_raw)}".lower()

    if _is_admin(client_raw) or "outscore meeting" in text or "sbc meeting" in text:
        return "admin"
    if any(token in text for token in ("bookkeeping", "bookeeping", "bank rec", "reconciliation", "financials", "books")):
        return "bookkeeping"
    if any(token in text for token in ("payroll", "941", "w2", "tax deposit", "adp")):
        return "payroll"
    if any(token in text for token in ("1040", "1065", "1120", "990", "tax return", "t/r", "return", "1099")):
        return "tax"
    if any(token in text for token in ("cfo", "forecast", "cash flow", "kpi", "advisory")):
        return "advisory"

    return "other"


def _parse_laura(file_path: Path, worksheet) -> list[TimeEntry]:
    entries: list[TimeEntry] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        client_task, staff, hours, entry_date = row[:4]
        if not _has_value(client_task) or not _has_value(hours) or not isinstance(entry_date, datetime):
            continue

        staff_name = _staff_name(staff) or "Laura"
        client_raw, task_raw = _split_client_task(client_task)
        entries.append(
            _make_entry(
                staff_name=staff_name,
                entry_date=entry_date.date(),
                client_raw=client_raw,
                task_raw=task_raw,
                hours=hours,
                file_path=file_path,
                sheet_name=worksheet.title,
                row_number=row_number,
            )
        )

    return entries


def _parse_wama(file_path: Path, worksheet) -> list[TimeEntry]:
    entries: list[TimeEntry] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        client, task, hours, entry_date, _total, paid = _pad_row(row, 6)[:6]
        if not _has_value(client) or not _has_value(hours):
            continue
        if _paid_by_upwork(paid):
            continue

        parsed_date = _coerce_date(entry_date)
        if parsed_date is None or parsed_date <= WAMA_VALID_AFTER:
            continue

        entries.append(
            _make_entry(
                staff_name="Wama",
                entry_date=parsed_date,
                client_raw=_clean_text(client),
                task_raw=_clean_text(task),
                hours=hours,
                file_path=file_path,
                sheet_name=worksheet.title,
                row_number=row_number,
                paid_marker=_clean_text(paid) or None,
            )
        )

    return entries


def _parse_beth(file_path: Path, worksheet) -> list[TimeEntry]:
    entries: list[TimeEntry] = []
    current_date: date | None = None

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        padded = _pad_row(row, 8)
        raw_date, client, task_one, task_two, task_three, hours = padded[:6]

        parsed_date = _coerce_date(raw_date)
        if parsed_date is not None:
            current_date = _repair_beth_date(parsed_date, file_path.name)

        if current_date is None or not _has_value(client) or not _has_value(hours):
            continue
        if _is_summary_row(client) or _is_summary_row(task_one) or _is_summary_row(task_two):
            continue

        task_raw = " ".join(_clean_text(value) for value in (task_one, task_two, task_three) if _has_value(value))
        entries.append(
            _make_entry(
                staff_name="Beth",
                entry_date=current_date,
                client_raw=_clean_text(client),
                task_raw=task_raw,
                hours=hours,
                file_path=file_path,
                sheet_name=worksheet.title,
                row_number=row_number,
            )
        )

    return entries


def _parse_julie(file_path: Path, worksheet) -> list[TimeEntry]:
    entries: list[TimeEntry] = []
    staff_name = "Julie"
    date_row = list(next(worksheet.iter_rows(min_row=13, max_row=13, values_only=True)))
    date_by_col = {index: _coerce_date(value) for index, value in enumerate(date_row)}

    for row_number, row in enumerate(worksheet.iter_rows(min_row=14, values_only=True), start=14):
        client = _clean_text(row[1] if len(row) > 1 else None)
        if not client or client.lower().startswith("total"):
            continue

        for index, value in enumerate(row):
            entry_date = date_by_col.get(index)
            if entry_date is None or not _is_positive_number(value):
                continue

            entries.append(
                _make_entry(
                    staff_name=staff_name,
                    entry_date=entry_date,
                    client_raw=client,
                    task_raw="",
                    hours=value,
                    file_path=file_path,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )

    return entries


def _make_entry(
    *,
    staff_name: str,
    entry_date: date,
    client_raw: str,
    task_raw: str,
    hours: object,
    file_path: Path,
    sheet_name: str,
    row_number: int,
    paid_marker: str | None = None,
) -> TimeEntry:
    parsed_hours = round(float(hours), 2)
    hourly_rate = STAFF_RATES[staff_name]
    service_type = infer_service_type(task_raw, client_raw)
    is_admin = service_type == "admin"

    return TimeEntry(
        staff_name=staff_name,
        entry_date=entry_date,
        client_raw=client_raw,
        task_raw=task_raw,
        hours=parsed_hours,
        hourly_rate=hourly_rate,
        labor_cost=round(parsed_hours * hourly_rate, 2),
        service_type=service_type,
        is_admin=is_admin,
        source_file=file_path.name,
        source_sheet=sheet_name,
        source_row=row_number,
        paid_marker=paid_marker,
    )


def _split_client_task(value: object) -> tuple[str, str]:
    text = _clean_text(value)
    if " - " not in text:
        return text, ""

    client, task = text.split(" - ", 1)
    return _clean_text(client), _clean_text(task)


def _repair_beth_date(value: date, filename: str) -> date:
    lower_name = filename.lower()
    if "dec" not in lower_name or "jan" not in lower_name:
        return value

    if value.month == 12 and value.year == 2026:
        return value.replace(year=2025)
    if value.month == 1 and value.year == 2025:
        return value.replace(year=2026)

    return value


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _staff_name(value: object) -> str:
    text = _clean_text(value)
    return text.split()[0].title() if text else ""


def _is_admin(value: object) -> bool:
    text = _clean_text(value).lower()
    return any(pattern in text for pattern in ADMIN_PATTERNS)


def _is_summary_row(value: object) -> bool:
    text = _clean_text(value).lower()
    return text.startswith("total") or text in {"billable hours", "admin hours"}


def _paid_by_upwork(value: object) -> bool:
    return "upwork" in _clean_text(value).lower()


def _has_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _is_positive_number(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return value > 0
    return False


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", " ", text).strip()


def _normalized_row(row: tuple[object, ...]) -> list[str]:
    return [_clean_text(value).lower() for value in row]


def _pad_row(row: tuple[object, ...], size: int) -> tuple[object, ...]:
    values = tuple(row)
    if len(values) >= size:
        return values
    return values + (None,) * (size - len(values))
